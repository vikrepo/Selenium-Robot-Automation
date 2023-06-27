import openpyxl
wk=openpyxl.load_workbook("C://Users//vikramadityaanand//Desktop//main1//Test-robot-modified.xlsx")

def fetch_number_of_rows(Sheetname):
    sh=wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname,row,cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value


print(fetch_cell_data("Sheet 1",1,2))